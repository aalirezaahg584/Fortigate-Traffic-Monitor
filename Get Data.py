import paramiko
import requests
import re
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from datetime import datetime
import os
import time

# ================================
# Fortigate Settings
# ================================
FG_IP     = "FW IP Address"
FG_USER   = "username with superadmin priv"
FG_PASS   = "password"
FG_PORT   = "443"
API_TOKEN = "API Token"
BASE_URL  = f"https://{FG_IP}:{FG_PORT}/api/v2"
HEADERS   = {"Authorization": f"Bearer {API_TOKEN}"}

POLICY_IDS = [ Policy IDs ]

STATE_FILE = "last_usage.json"
requests.packages.urllib3.disable_warnings()

# ============================================
# API → Fetch Policy Name
# ============================================
def get_policy_name(policy_id):
    url = f"{BASE_URL}/cmdb/firewall/policy/{policy_id}"
    r = requests.get(url, headers=HEADERS, verify=False)
    if r.status_code != 200:
        return "Unknown"
    data = r.json()
    if "results" in data and len(data["results"]) > 0:
        return data["results"][0].get("name", "NO-NAME")
    return "Unknown"

# ============================================
# SSH helper
# ============================================
def ssh_send(shell, cmd, wait=0.4):
    shell.send(cmd + "\n")
    time.sleep(wait)
    return shell.recv(99999).decode(errors="ignore")

def to_gb(x):
    return round(x / (1024**3), 2)

# ============================================
# Load previous state
# ============================================
if os.path.exists(STATE_FILE):
    last_usage = json.load(open(STATE_FILE))
else:
    last_usage = {}

# ============================================
# SSH CONNECT
# ============================================
client = paramiko.SSHClient()
client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
client.connect(FG_IP, username=FG_USER, password=FG_PASS)

shell = client.invoke_shell()
time.sleep(1)
shell.recv(99999)

ssh_send(shell, "config vdom")
ssh_send(shell, "edit root")

# ============================================
# Collect Data from Fortigate
# ============================================
collected = []

for pid in POLICY_IDS:
    name = get_policy_name(pid)
    out = ssh_send(shell, f"diagnose firewall iprope show 00100004 {pid}", wait=0.6)
    if "idx:" not in out:
        continue
    m = re.search(r"idx:(\d+).*?bytes:(\d+)", out, re.S)
    if not m:
        continue
    idx = m.group(1)
    bytes_now = int(m.group(2))
    bytes_last = last_usage.get(idx, 0)
    collected.append({
        "id": pid,
        "name": name,
        "total_gb": to_gb(bytes_now)
    })
    last_usage[idx] = bytes_now

client.close()

# ============================================
#  Split  Primary & Secondary For Some Scenario
# ============================================
secondary_ids = [63, 208]
primary_data = [x for x in collected if x["id"] not in secondary_ids]
secondary_data = [x for x in collected if x["id"] in secondary_ids]

# ============================================
# Excel Output
# ============================================
today = datetime.now().strftime("%Y-%m-%d")
excel_name = f"policy-traffic-{today}.xlsx"

wb = Workbook()

def create_sheet(wb, sheet_name, data):
    ws = wb.create_sheet(title=sheet_name)

    #  Sequence Data  
    data_sorted = sorted(data, key=lambda x: x['total_gb'], reverse=True)

    # Header
    headers = ["Policy ID", "Policy Name", "Total Traffic (GB)"]
    ws.append(headers)

    # Styles
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Conditional color
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    total_sum = 0
    for item in data_sorted:
        row = [item["id"], item["name"], item["total_gb"]]
        ws.append(row)
        for col_idx in range(1, 4):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.alignment = center_align
            cell.border = thin_border
        traffic_cell = ws.cell(row=ws.max_row, column=3)
        if item["total_gb"] <= 2:
            traffic_cell.fill = green_fill
        elif 2 < item["total_gb"] <= 4:
            traffic_cell.fill = yellow_fill
        else:
            traffic_cell.fill = red_fill
        total_sum += item["total_gb"]

    # total sum 
    ws.append(["", "Total", round(total_sum, 2)])
    for col_idx in range(1, 4):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.alignment = center_align
        cell.border = thin_border
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=3).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 5

    # create chart 
    chart = BarChart()
    chart.title = f"Traffic per Policy (GB) - {sheet_name}"
    chart.x_axis.title = "Policy Name"
    chart.y_axis.title = "Total Traffic (GB)"
    chart.height = 15
    chart.width = 25

    data_ref = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row-1)
    categories_ref = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)
    chart.style = 9
    ws.add_chart(chart, "E2")

    return total_sum

# delete default excel sheet  
wb.remove(wb.active)

# create sheet 
primary_total = create_sheet(wb, "Primary", primary_data)
secondary_total = create_sheet(wb, "Secondary", secondary_data)
overall_total = primary_total + secondary_total

# ============================================
# create  Total sheet
# ============================================
ws_total = wb.create_sheet(title="Total")

headers_total = ["Metric", "Value"]
ws_total.append(headers_total)

# Styles
center_align = Alignment(horizontal="center", vertical="center")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for col_idx, header in enumerate(headers_total, start=1):
    cell = ws_total.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

#  Total Data
metrics = [
    ["Report Date", today],
    ["Total Policies", len(collected)],
    ["Primary Policies", len(primary_data)],
    ["Secondary Policies", len(secondary_data)],
    ["Primary Traffic (GB)", round(primary_total,2)],
    ["Secondary Traffic (GB)", round(secondary_total,2)],
    ["Overall Traffic (GB)", round(overall_total,2)],
]

#  report max & min
if collected:
    max_item = max(collected, key=lambda x: x["total_gb"])
    min_item = min(collected, key=lambda x: x["total_gb"])
    metrics.append(["Max Traffic Policy", f"{max_item['name']} ({max_item['total_gb']} GB)"])
    metrics.append(["Min Traffic Policy", f"{min_item['name']} ({min_item['total_gb']} GB)"])

for metric in metrics:
    ws_total.append(metric)
    for col_idx in range(1,3):
        cell = ws_total.cell(row=ws_total.max_row, column=col_idx)
        cell.alignment = center_align
        cell.border = thin_border

# Auto-fit columns
for col in ws_total.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            value = str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        except:
            pass
    ws_total.column_dimensions[column].width = max_length + 5

#  Primary و Secondary
pie = PieChart()
pie.title = "Traffic Distribution"
labels = Reference(ws_total, min_col=1, min_row=6, max_row=7)  # Primary و Secondary
data = Reference(ws_total, min_col=2, min_row=6, max_row=7)
pie.add_data(data, titles_from_data=False)
pie.set_categories(labels)
ws_total.add_chart(pie, "E2")

# Save Done 
wb.save(excel_name)
json.dump(last_usage, open(STATE_FILE, "w"), indent=4)

print("DONE → Excel Created with Primary, Secondary & Total Sheets:", excel_name)
